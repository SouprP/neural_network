import numpy as np
import os
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook

class NeuralNetworkAckley:
    def __init__(self, input_size: int, hidden_size1: int, hidden_size2: int, output_size: int,
                 activation_function='relu', filename=None, debug=False, iters_check=100):
        self.input_size = input_size
        self.hidden_size1 = hidden_size1
        self.hidden_size2 = hidden_size2
        self.output_size = output_size
        self.activation_function = activation_function
        self.filename = filename
        self.debug = debug


        self.iters_check = iters_check
        self.start_iteration = 0
        self.current_iteration = 0

        self.alpha = None
        self.iterations = None
        self.mse_history = []  # Przechowywanie MSE na potrzeby wizualizacji

        # Tworzenie katalogu na dane
        if not os.path.exists("./data/"):
            os.mkdir("./data/")

        # Wczytywanie wag lub inicjalizacja
        if self.filename and os.path.exists(f"./data/{self.filename}.npz"):
            self.load_data()
        else:
            self.W1, self.b1, self.W2, self.b2, self.W3, self.b3 = self.init_params()

    def init_params(self):
        W1 = np.random.rand(self.hidden_size1, self.input_size) - 0.5
        b1 = np.random.rand(self.hidden_size1, 1) - 0.5
        W2 = np.random.rand(self.hidden_size2, self.hidden_size1) - 0.5
        b2 = np.random.rand(self.hidden_size2, 1) - 0.5
        W3 = np.random.rand(self.output_size, self.hidden_size2) - 0.5
        b3 = np.random.rand(self.output_size, 1) - 0.5
        return W1, b1, W2, b2, W3, b3

    def load_data(self):
        data = np.load(f"./data/{self.filename}.npz")
        self.W1 = data['W1']
        self.b1 = data['b1']
        self.W2 = data['W2']
        self.b2 = data['b2']
        self.W3 = data['W3']
        self.b3 = data['b3']
        self.start_iteration = data['last_iteration']

        # print(self.start_iteration)

        if self.debug:
            print(f"Model loaded from ./data/{self.filename}.npz")

    def save_data(self):
        if self.filename == None:
            return

        np.savez("./data/" + self.filename, W1=self.W1, b1=self.b1, 
                 W2=self.W2, b2=self.b2, W3=self.W3, b3=self.b3,
                 last_iteration=self.current_iteration)

        if self.debug == True:
            print(f"Model saved to ./data/{self.filename}.npz")

    def relu(self, x):
        return np.maximum(0, x)

    def relu_deriv(self, x):
        return x > 0

    def sigmoid(self, x):
        return 1 / (1 + np.exp(-x))

    def sigmoid_deriv(self, x):
        sig = self.sigmoid(x)
        return sig * (1 - sig)

    def forward_prop(self, X):
        Z1 = self.W1.dot(X) + self.b1
        A1 = self.relu(Z1) if self.activation_function == 'relu' else self.sigmoid(Z1)
        Z2 = self.W2.dot(A1) + self.b2
        A2 = self.relu(Z2) if self.activation_function == 'relu' else self.sigmoid(Z2)
        Z3 = self.W3.dot(A2) + self.b3
        return Z1, A1, Z2, A2, Z3

    def backward_prop(self, Z1, A1, Z2, A2, Z3, X, Y):
        dZ3 = Z3 - Y
        dW3 = 1 / X.shape[1] * dZ3.dot(A2.T)
        db3 = 1 / X.shape[1] * np.sum(dZ3, axis=1, keepdims=True)

        dZ2 = self.W3.T.dot(dZ3) * (self.relu_deriv(Z2) if self.activation_function == 'relu' else self.sigmoid_deriv(Z2))
        dW2 = 1 / X.shape[1] * dZ2.dot(A1.T)
        db2 = 1 / X.shape[1] * np.sum(dZ2, axis=1, keepdims=True)

        dZ1 = self.W2.T.dot(dZ2) * (self.relu_deriv(Z1) if self.activation_function == 'relu' else self.sigmoid_deriv(Z1))
        dW1 = 1 / X.shape[1] * dZ1.dot(X.T)
        db1 = 1 / X.shape[1] * np.sum(dZ1, axis=1, keepdims=True)

        return dW1, db1, dW2, db2, dW3, db3

    def update_params(self, dW1, db1, dW2, db2, dW3, db3):
        self.W1 -= self.alpha * dW1
        self.b1 -= self.alpha * db1
        self.W2 -= self.alpha * dW2
        self.b2 -= self.alpha * db2
        self.W3 -= self.alpha * dW3
        self.b3 -= self.alpha * db3

    def get_mse(self, predictions, actual):
        return np.mean((predictions - actual) ** 2)
    
    def get_msa(self, predictions, actual):
        return np.mean((predictions * actual) ** 2)

    def train(self, X, Y, alpha, iterations, sheet=None):
        self.alpha = alpha
        self.iterations = iterations

        self.current_iteration = self.start_iteration
        for i in range(iterations):
            iteration_number = self.current_iteration + i

            Z1, A1, Z2, A2, Z3 = self.forward_prop(X)
            dW1, db1, dW2, db2, dW3, db3 = self.backward_prop(Z1, A1, Z2, A2, Z3, X, Y)
            self.update_params(dW1, db1, dW2, db2, dW3, db3)

            if i % self.iters_check == 0:
                mse = self.get_mse(Z3, Y)
                self.mse_history.append((i, mse))
                print(f"Iteration {iteration_number}: MSE {mse:.4f}")
                if sheet:
                    sheet.append([i, mse])

        self.current_iteration = iteration_number

        self.save_data()

    def train_multiple(self, X, Y, alpha, iterations, excel_filename="training_data.xlsx", start_col=1):
        """
        Trains the neural network for multiple runs and saves metrics to Excel.

        Parameters:
            X (numpy array): Input data.
            Y (numpy array): True labels.
            alpha (float): Learning rate for gradient descent.
            iterations (int): Number of training iterations.
            excel_filename (str): Name of the Excel file to save metrics.
            start_col (int): Column number where the training output should begin.

        Returns:
            None
        """
        self.alpha = alpha
        self.iterations = iterations

        # load / create excel file
        if os.path.exists(excel_filename):
            workbook = load_workbook(excel_filename)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active

        # metadata init
        if sheet.cell(row=1, column=start_col).value is None:
            sheet.cell(row=1, column=start_col, value="Layer 1")
            sheet.cell(row=1, column=start_col + 1, value="Layer 2")
            sheet.cell(row=1, column=start_col + 2, value="Alpha")
            sheet.cell(row=2, column=start_col, value=self.hidden_size1)
            sheet.cell(row=2, column=start_col + 1, value=self.hidden_size2)
            sheet.cell(row=2, column=start_col + 2, value=alpha)
            sheet.cell(row=4, column=start_col, value="Iteration")
            sheet.cell(row=4, column=start_col + 1, value="MSE")
            # sheet.cell(row=4, column=start_col + 2, value="MSA")

        # get last iter from .npz file
        self.current_iteration = self.start_iteration
        iteration_number = 0

        # where to write values, last non empty cell in column
        row_to_write = 5
        while sheet.cell(row=row_to_write, column=start_col).value is not None:
            row_to_write += 1

        for i in range(1, iterations):
            iteration_number = self.current_iteration + i

            Z1, A1, Z2, A2, Z3 = self.forward_prop(X)
            dW1, db1, dW2, db2, dW3, db3 = self.backward_prop(Z1, A1, Z2, A2, Z3, X, Y)

            # update weights
            self.update_params(dW1, db1, dW2, db2, dW3, db3)

            # calculate MSE
            mse = self.get_mse(Z3, Y)
            # msa = self.get_msa(Z3, Y)

            # save metrics
            if i % self.iters_check == 0:
                sheet.cell(row=row_to_write, column=start_col, value=self.current_iteration + i)
                sheet.cell(row=row_to_write, column=start_col + 1, value=mse)
                # sheet.cell(row=row_to_write, column=start_col + 2, value=msa)

                print(f"Iteration {iteration_number}: MSE {mse:.6f}")
                row_to_write += 1

        # save excel file
        self.current_iteration = iteration_number
        workbook.save(excel_filename)

        # print(f"Training metrics appended to {excel_filename}.")


    def predict(self, X):
        _, _, _, _, Z3 = self.forward_prop(X)
        return Z3


    def plot_mse(self):
        iterations, mse_values = zip(*self.mse_history)
        plt.plot(iterations, mse_values)
        plt.xlabel("Iteration")
        plt.ylabel("MSE")
        plt.show()
