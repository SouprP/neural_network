import numpy as np
import os.path
import matplotlib.pyplot as plt
# from openpyxl import Workbook
from openpyxl import Workbook, load_workbook

class NeuralNetworkMnist:
    def __init__(self, input_size : int,
                 hidden_size1 : int, hidden_size2 : int, output_size : int, activation_function='relu',
                 filename = None, debug=False, iters_check=10):
        """
        Initializes the nueral network with the given layer sizes.
        """

        self.input_size = input_size
        self.hidden_size1 = hidden_size1
        self.hidden_size2 = hidden_size2
        self.output_size = output_size
        self.activation_function = activation_function

        # alpha and iteration count is set later in train()
        self.alpha = None
        self.iterations = None

        # debug and data loading
        self.filename = filename
        self.debug = debug
        self.iters_check = iters_check
        self.start_iteration = 0
        self.current_iteration = 0

        # plotting data
        self.accuracy_data = np.empty((0, 2))
        #print(self.accuracy_data.shape)

        # setup weigths and biases if data file doesnt exists (or is not needed)
        if not os.path.exists("./data/"):
            os.mkdir("./data/")

        if os.path.exists(f"./data/{self.filename}.npz") == False:
            self.W1, self.b1, self.W2, self.b2, self.W3, self.b3 = self.init_params()
        else:
            self.load_data()



    def init_params(self):
        """
        Initializes weights and biases with small random values.
        """
        W1 = np.random.rand(self.hidden_size1, self.input_size) - 0.5
        b1 = np.random.rand(self.hidden_size1, 1) - 0.5
        W2 = np.random.rand(self.hidden_size2, self.hidden_size1) - 0.5
        b2 = np.random.rand(self.hidden_size2, 1) - 0.5
        W3 = np.random.rand(self.output_size, self.hidden_size2) - 0.5
        b3 = np.random.rand(self.output_size, 1) - 0.5
        return W1, b1, W2, b2, W3, b3

    def load_data(self):
        data = np.load("./data/" + self.filename + ".npz")

        self.W1 = data['W1']
        self.b1 = data['b1']
        self.W2 = data['W2']
        self.b2 = data['b2']
        self.W3 = data['W3']
        self.b3 = data['b3']
        self.start_iteration = data['last_iteration']

        # print("Iteration start: ", self.start_iteration)

        if self.debug == True:
            print(f"Model loaded from ./data/{self.filename}.npz")

    def save_data(self):
        if self.filename == None:
            return

        np.savez("./data/" + self.filename, W1=self.W1, b1=self.b1, 
                 W2=self.W2, b2=self.b2, W3=self.W3, b3=self.b3,
                 last_iteration=self.current_iteration)

        if self.debug == True:
            print(f"Model saved to ./data/{self.filename}.npz")

    def relu(self, x):
        """
        ReLU activation function, beetwen -INF to 0 it's value is 0,
        from 0 to INF it's value is max(0, x).
        """
        return np.maximum(0, x)

    def relu_deriv(self, x):
        """
        Derivative of the ReLU activation function.
        """
        return x > 0

    def sigmoid(self, x):
        return 1 / (1 + np.exp(-x))

    def sigmoid_deriv(self, x):
        sig = self.sigmoid(x)
        return sig * (1 - sig)

    def softmax(self, x):
        """
        Softmax activation function for multi-class probability distribution.
        """
        expX = np.exp(x - np.max(x, axis=0, keepdims=True))
        return expX / np.sum(expX, axis=0, keepdims=True)

    def forward_prop(self, x):
        """
        Performs forward propagation through the network.
        Calculates activations for both the hidden and output layers.

        Parameters:
            X (numpy array): Input data of shape (input_size, number of samples).

        Returns:
            Z1, A1, Z2, A2 (numpy arrays): Intermediate values:
                Z1 - Weighted sum for the hidden layer.
                A1 - Activation after applying ReLU to Z1.
                Z2 - Weighted sum for the output layer.
                A2 - Activation after applying softmax to Z2 (final output probabilities).
        """
        # First hidden layer
        Z1 = self.W1.dot(x) + self.b1
        if self.activation_function == 'relu':
            A1 = self.relu(Z1)
        elif self.activation_function == 'sigmoid':
            A1 = self.sigmoid(Z1)
        else:
            raise ValueError("Invalid activation function. Choose 'relu' or 'sigmoid'.")

        # Second hidden layer
        Z2 = self.W2.dot(A1) + self.b2
        if self.activation_function == 'relu':
            A2 = self.relu(Z2)
        elif self.activation_function == 'sigmoid':
            A2 = self.sigmoid(Z2)
        else:
            raise ValueError("Invalid activation function. Choose 'relu' or 'sigmoid'.")

        # Output layer (always uses softmax for classification)
        Z3 = self.W3.dot(A2) + self.b3
        A3 = self.softmax(Z3)

        return Z1, A1, Z2, A2, Z3, A3

    def one_hot(self, Y, num_classes):
        """
        Converts labels into one-hot encoding.
        https://en.wikipedia.org/wiki/One-hot
        """
        one_hot_Y = np.zeros((num_classes, Y.size))
        one_hot_Y[Y, np.arange(Y.size)] = 1
        return one_hot_Y

    def backward_prop(self, Z1, A1, Z2, A2, Z3, A3, X, Y):
        """
        Performs backward propagation to compute gradients for the weights and biases.

        Parameters:
            Z1, A1 (numpy arrays): Values from the hidden layer.
            Z2, A2 (numpy arrays): Values from the output layer.
            X (numpy array): Input data
            Y (numpy array): Output labels

        Returns:
            dW1, db1, dW2, db2 (numpy arrays): Gradients for weights and biases.
        """

        # One-hot encoding
        one_hot_Y = self.one_hot(Y, self.output_size)

        # Gradients for output layer
        dZ3 = A3 - one_hot_Y
        dW3 = 1 / X.shape[1] * dZ3.dot(A2.T)
        db3 = 1 / X.shape[1] * np.sum(dZ3, axis=1, keepdims=True)

        # Gradients for second hidden layer
        dZ2 = self.W3.T.dot(dZ3) * (self.relu_deriv(Z2) if self.activation_function == 'relu' else self.sigmoid_deriv(Z2))
        dW2 = 1 / X.shape[1] * dZ2.dot(A1.T)
        db2 = 1 / X.shape[1] * np.sum(dZ2, axis=1, keepdims=True)

        # Gradients for first hidden layer
        dZ1 = self.W2.T.dot(dZ2) * (self.relu_deriv(Z1) if self.activation_function == 'relu' else self.sigmoid_deriv(Z1))
        dW1 = 1 / X.shape[1] * dZ1.dot(X.T)
        db1 = 1 / X.shape[1] * np.sum(dZ1, axis=1, keepdims=True)

        return dW1, db1, dW2, db2, dW3, db3

    def update_params(self, dW1, db1, dW2, db2, dW3, db3, alpha):
        """
        Updates weights and biases using gradient descent.
        """
        self.W1 -= alpha * dW1
        self.b1 -= alpha * db1
        self.W2 -= alpha * dW2
        self.b2 -= alpha * db2
        self.W3 -= alpha * dW3
        self.b3 -= alpha * db3

    def get_predictions(self, A3):
        """
        Returns the predicted classes.
        """
        return np.argmax(A3, axis=0)

    def get_accuracy(self, predictions, Y):
        """
        Computes the accuracy of predictions.
        """
        return np.sum(predictions == Y) / Y.size

    def train(self, X, Y, alpha, iterations, excel_filename="training_data.xlsx"): 
        """
        Trains the neural network and saves metrics to an Excel file.

        Parameters:
            X (numpy array): Input data.
            Y (numpy array): True labels.
            alpha (float): Learning rate for gradient descent.
            iterations (int): Number of training iterations.
            excel_filename (str): Name of the Excel file to save metrics.

        Returns:
            None
        """
        self.alpha = alpha
        self.iterations = iterations

        # load / create
        if os.path.exists(excel_filename):
            workbook = load_workbook(excel_filename)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active

            # write metadata in the first rows
            sheet.append(["Hidden Layer 1 Size", "Hidden Layer 2 Size", "Learning Rate"])
            sheet.append([self.hidden_size1, self.hidden_size2, alpha])
            sheet.append([])
            sheet.append(["Iteration", "Accuracy", "MSE"])

        # from what iteration to start
        self.current_iteration = self.start_iteration
        iteration_number = 0

        for i in range(iterations):
            iteration_number = self.current_iteration + i

            Z1, A1, Z2, A2, Z3, A3 = self.forward_prop(X)
            dW1, db1, dW2, db2, dW3, db3 = self.backward_prop(Z1, A1, Z2, A2, Z3, A3, X, Y)

            # update weights
            self.update_params(dW1, db1, dW2, db2, dW3, db3, alpha)

            # compute accuracy and MSE
            predictions = self.get_predictions(A3)
            accuracy = self.get_accuracy(predictions, Y)
            mse = np.mean((A3 - self.one_hot(Y, self.output_size)) ** 2)

            if i % self.iters_check == 0:
                sheet.append([iteration_number, accuracy, mse])
                print(f"Iteration {iteration_number}: Accuracy {accuracy:.4f}, MSE {mse:.6f}")

                self.accuracy_data = np.append(self.accuracy_data, [[iteration_number, accuracy]], axis=0)

        # save the excel file
        self.current_iteration = iteration_number
        workbook.save(excel_filename)
        #print(f"Training metrics saved to {excel_filename}")


    def train_multiple(self, X, Y, alpha, iterations, excel_filename="training_data.xlsx", start_col=1):
        """
        Trains the neural network for multiple runs and saves metrics to Excel,
        starting output from the specified column.

        Parameters:
            X (numpy array): Input data.
            Y (numpy array): True labels.
            alpha (float): Learning rate for gradient descent.
            iterations (int): Number of training iterations.
            excel_filename (str): Name of the Excel file to save metrics.
            start_col (int): Column number where the training output should begin.

        Returns:
            None
        """
        self.alpha = alpha
        self.iterations = iterations

        # load / create excel file
        if os.path.exists(excel_filename):
            workbook = load_workbook(excel_filename)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active

        # check if metadata exists at the top, if not, write it based on start_col
        if sheet.cell(row=1, column=start_col).value is None:
            # write metadata in the first row
            sheet.cell(row=1, column=start_col, value="Layer 1")
            sheet.cell(row=1, column=start_col + 1, value="Layer 2")
            sheet.cell(row=1, column=start_col + 2, value="Alpha")

            sheet.cell(row=2, column=start_col, value=self.hidden_size1)
            sheet.cell(row=2, column=start_col + 1, value=self.hidden_size2)
            sheet.cell(row=2, column=start_col + 2, value=alpha)

            sheet.cell(row=3, column=start_col, value="")
            
            sheet.cell(row=4, column=start_col, value="Iteration")
            sheet.cell(row=4, column=start_col + 1, value="Accuracy")
            sheet.cell(row=4, column=start_col + 2, value="MSE")
            # sheet.cell(row=4, column=start_col + 3, value="MSA")

        # get last iter from .npz file
        self.current_iteration = self.start_iteration
        iteration_number = 0

        # where to write values, last non empty cell in column
        row_to_write = 5  # Start writing from row 5 (skip metadata and headers)
        while sheet.cell(row=row_to_write, column=start_col).value is not None:
            row_to_write += 1

        for i in range(1, iterations):
            iteration_number = self.current_iteration + i

            Z1, A1, Z2, A2, Z3, A3 = self.forward_prop(X)
            dW1, db1, dW2, db2, dW3, db3 = self.backward_prop(Z1, A1, Z2, A2, Z3, A3, X, Y)

            # update weights
            self.update_params(dW1, db1, dW2, db2, dW3, db3, alpha)

            # compute accuracy and MSE
            predictions = self.get_predictions(A3)
            accuracy = self.get_accuracy(predictions, Y)
            mse = np.mean((A3 - self.one_hot(Y, self.output_size)) ** 2)
            # msa = np.mean(A3 ** 2)

            # save metrics
            if i % self.iters_check == 0:
                sheet.cell(row=row_to_write, column=start_col, value=iteration_number)
                sheet.cell(row=row_to_write, column=start_col + 1, value=accuracy)
                sheet.cell(row=row_to_write, column=start_col + 2, value=mse)
                # sheet.cell(row=row_to_write, column=start_col + 3, value=msa)

                # print(f"Iteration {iteration_number}: Accuracy {accuracy:.4f}, MSE {mse:.6f}")

                self.accuracy_data = np.append(self.accuracy_data, [[iteration_number, accuracy]], axis=0)

                row_to_write += 1

        # save the Excel file
        self.current_iteration = iteration_number
        workbook.save(excel_filename)
        # print(f"Training metrics saved to {excel_filename} starting at column {start_col}")

    def predict(self, X):
        """
        Makes predictions for input data (image for example).
        """
        _, _, _, _, _, A3 = self.forward_prop(X)
        return self.get_predictions(A3)

    def plot(self):
        x_points, y_points = np.split(self.accuracy_data, 2, axis=1)

        plt.plot(x_points, y_points)
        plt.title(f"Prediction accuracy, alpha={self.alpha}, iterations={self.iterations}")
        plt.show()